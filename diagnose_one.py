#!/usr/bin/env python3
"""
对比单条记录的 Excel 值 vs PDF 提取值，方便定位为什么不匹配。

用法：
    python diagnose_one.py --excel expected_data.xlsx --sheet "v1a check" --pdf-dir output --name "TAN AH POOK"
"""
import argparse
import re
from pathlib import Path

import fitz
from openpyxl import load_workbook

# 复用 verify_letters.py 里的逻辑
from verify_letters import (
    EXCEL_COLS, FIELD_KIND, BOLD_EXPECTED,
    safe_filename, read_pdf_text, extract_fields_from_pdf,
    get_amount_bold_map, norm_amount, normalize,
)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True, type=Path)
    p.add_argument("--sheet", default="v1a check")
    p.add_argument("--pdf-dir", required=True, type=Path)
    p.add_argument("--name", required=True, help="Name in POD 的值")
    args = p.parse_args()

    wb = load_workbook(args.excel, data_only=True)
    ws = wb[args.sheet]

    # 列号
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    header_map = {str(c.value).strip(): c.column for c in header_row if c.value}
    col_idx = {k: header_map[v] for k, v in EXCEL_COLS.items() if v in header_map}

    # 找匹配的那一行
    target = args.name.strip().upper()
    found_row = None
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx["name"] - 1]
        if cell.value and str(cell.value).strip().upper() == target:
            found_row = row
            break

    if not found_row:
        print(f"❌ 在 Excel 里找不到 Name in POD = {args.name!r}")
        return

    # PDF
    pdf_path = args.pdf_dir / f"{safe_filename(args.name)}.pdf"
    if not pdf_path.exists():
        print(f"❌ 找不到 PDF：{pdf_path}")
        return

    pdf_text = read_pdf_text(pdf_path)
    pdf_fields = extract_fields_from_pdf(pdf_text, args.name)
    bold_map = get_amount_bold_map(pdf_path)

    print(f"\n========== 对比 {args.name} ==========")
    print(f"PDF 文件：{pdf_path}\n")
    print(f"{'字段':<35} {'类型':<8} {'Excel 原值':<35} {'PDF 原值':<35} {'归一 Excel':<20} {'归一 PDF':<20} {'结果'}")
    print("-" * 200)
    for key, col_name in EXCEL_COLS.items():
        if key not in col_idx:
            print(f"{col_name:<35} (列不存在)")
            continue
        excel_val = found_row[col_idx[key] - 1].value
        pdf_val = pdf_fields.get(key, "")
        kind = FIELD_KIND[key]
        e_norm = normalize(excel_val, kind)
        p_norm = normalize(pdf_val, kind)
        match = "✓" if e_norm == p_norm else "✗"
        print(f"{col_name:<35} {kind:<8} {repr(excel_val)[:33]:<35} "
              f"{repr(pdf_val)[:33]:<35} {e_norm[:18]:<20} {p_norm[:18]:<20} {match}")

    print(f"\n========== Bold 检查 ==========")
    for key, expected in BOLD_EXPECTED.items():
        if key not in col_idx:
            continue
        pdf_val = pdf_fields.get(key, "")
        actual = bold_map.get(norm_amount(pdf_val))
        if actual is None:
            print(f"{EXCEL_COLS[key]:<35} 期望={'粗体' if expected else '非粗体'}, 实际=找不到")
        else:
            mark = "✓" if actual == expected else "✗"
            print(f"{EXCEL_COLS[key]:<35} 期望={'粗体' if expected else '非粗体'}, "
                  f"实际={'粗体' if actual else '非粗体'}  {mark}")

    print(f"\n========== bold_map (PDF 里所有 SGD 金额) ==========")
    for amt, is_bold in bold_map.items():
        print(f"  SGD {amt:<15} {'粗体' if is_bold else '非粗体'}")


if __name__ == "__main__":
    main()
