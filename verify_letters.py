#!/usr/bin/env python3
"""
对比每封 PDF 信件的内容与 Excel 表里的数据。

匹配方式：PDF 文件名 = Excel "Name in POD" 列（空格转下划线后）
输出：把 Excel 复制一份，不匹配的单元格高亮成红色，并加一列 Status

用法：
    python verify_letters.py --excel expected_data.xlsx --pdf-dir output/ --out report.xlsx
    python verify_letters.py --excel expected_data.xlsx --sheet expected_data --pdf-dir output/
"""
from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import load_workbook


# Excel 列名 -> 内部 key（如果列名不一样，改这里）
EXCEL_COLS = {
    "name": "Name in POD",
    "addr1": "Vlookup POD address 1",
    "addr2": "Vlookup POD address 2",
    "addr3": "Vlookup POD address 3",
    "country": "Country",
    "pod_date": "Revised POD Date w/o formula",
    "pod_amount": "POD amount (2dp)",
    "first_div": "First Dividend amount (2 dp)",
    "admitted": "Admitted amount (2dp)",
    "final_div": "Final Dividend amount (2dp)",
    "bank_name": "Bank Name",
    "bank_account": "Bank Account Number",
}

FIELD_KIND = {
    "name": "text",
    "addr1": "text",
    "addr2": "text",
    "addr3": "digits",   # 邮编
    "country": "text",
    "pod_date": "date",
    "pod_amount": "amount",
    "first_div": "amount",
    "admitted": "amount",
    "final_div": "amount",
    "bank_name": "text",
    "bank_account": "digits",
}

# 期望的粗体状态：True = SGD 应粗体；False = 不应粗体；不在表里则不检查粗体
# 只有 Final Dividend 那行的 "SGD XXXX" 是粗体
BOLD_EXPECTED = {
    "pod_amount": False,
    "first_div": False,
    "admitted": False,
    "final_div": True,
}

# 问题摘要里用的简称（出现在最后一列）
ISSUE_LABEL = {
    "name": "Name in POD",
    "addr1": "POD address 1",
    "addr2": "POD address 2",
    "addr3": "POD address 3",
    "country": "Country",
    "pod_date": "POD date",
    "pod_amount": "POD amount",
    "first_div": "First Dividend",
    "admitted": "Admitted amount",
    "final_div": "Final Dividend",
    "bank_name": "Bank Name",
    "bank_account": "Bank Account Number",
}


# ---------- 归一化 ----------

def norm_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().upper()


def norm_digits(v) -> str:
    """归一化数字字符串：去掉非数字 + 去掉前导零。
    这样 Excel 存 85460455 vs PDF 读到 085460455 能匹配。
    注意：Excel 占位 "0" 不再当空看待；地址里的 "0" 由专门的异常检测处理。
    """
    if v is None:
        return ""
    s = re.sub(r"\D", "", str(v))
    return s.lstrip("0") or ("0" if s else "")


def norm_amount(v) -> str:
    if v is None or v == "":
        return ""
    s = str(v).upper().replace("SGD", "").replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return f"{float(m.group()):.2f}" if m else ""


MONTHS = {
    "JAN": 1, "JANUARY": 1, "FEB": 2, "FEBRUARY": 2, "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4, "MAY": 5, "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7, "AUG": 8, "AUGUST": 8, "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10, "NOV": 11, "NOVEMBER": 11, "DEC": 12, "DECEMBER": 12,
}


def norm_date(v) -> str:
    """归一化成 'YYYY-MM-DD' 字符串。"""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().upper()
    # 例如 "22 FEBRUARY 2022" 或 "22-FEB-2022" 或 "22/02/2022"
    m = re.match(r"(\d{1,2})[\s\-/]+([A-Z]+|\d{1,2})[\s\-/]+(\d{4})", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if mo.isdigit():
            month = int(mo)
        else:
            month = MONTHS.get(mo, 0)
        if month:
            return f"{int(y):04d}-{month:02d}-{int(d):02d}"
    # ISO 格式 YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return s


def normalize(v, kind: str) -> str:
    if kind == "amount":
        return norm_amount(v)
    if kind == "date":
        return norm_date(v)
    if kind == "digits":
        return norm_digits(v)
    return norm_text(v)


# ---------- PDF 解析 ----------

def safe_filename(name) -> str:
    cleaned = re.sub(r"[^\w\s\-]", "", str(name)).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "UNKNOWN"


def read_pdf_text(pdf_path: Path) -> str:
    doc = fitz.open(str(pdf_path))
    text = "\n".join(page.get_text("text") for page in doc)
    doc.close()
    return text


# 抬头/页脚关键字，匹配则跳过（用于从 PDF 提取收信人名字）
_NAME_SKIP_PATTERNS = [
    r"SUNSHINE\s+EMPIRE",
    r"COMPULSORY\s+LIQUIDATION",
    r"Co\.?\s*Reg",
    r"PricewaterhouseCoopers",
    r"Straits\s+View",
    r"East\s+Tower",
    r"Marina\s+One",
    r"Singapore\s+\d",
    r"Telephone",
    r"Facsimile",
    r"PTE\.?\s*LTD",
    r"The\s+affairs",
    r"The\s+Liquidators",
    r"NOTICE\s+OF",
    r"Dear\s+Sir",
    r"Our\s+Ref",
]
_NAME_SKIP_RE = re.compile("|".join(_NAME_SKIP_PATTERNS), re.IGNORECASE)


def extract_recipient_name_from_text(text: str) -> str:
    """从 PDF 文本里找收信人名字（地址块的第一行）。"""
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if _NAME_SKIP_RE.search(line):
            continue
        if re.search(r"\d", line):
            continue
        line = re.sub(r"\s+", " ", line)
        # 全大写字母组成（允许空格、点、连字符、撇号、斜杠、括号、&）
        if re.fullmatch(r"[A-Z][A-Z\s\.\-'/()&]{1,}", line):
            if len(re.findall(r"[A-Z]", line)) >= 2:
                return line
    return ""


def _is_span_bold(span: dict) -> bool:
    """判断 PyMuPDF 的 span 是否粗体。"""
    flags = span.get("flags", 0)
    if flags & 16:  # PyMuPDF flag bit 4 = bold
        return True
    font = span.get("font", "").lower()
    return any(tag in font for tag in ("bold", "black", "heavy"))


def get_amount_bold_map(pdf_path: Path) -> dict[str, bool]:
    """扫 PDF 里所有 'SGD XXX' 形式的金额，返回 {归一化金额: 是否粗体}。

    同一金额出现多次时，只要任意一次是粗体就记为粗体。
    """
    result: dict[str, bool] = {}
    doc = fitz.open(str(pdf_path))
    try:
        for page in doc:
            data = page.get_text("dict")
            for block in data.get("blocks", []):
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    spans = line.get("spans", [])
                    if not spans:
                        continue
                    line_text = "".join(s.get("text", "") for s in spans)
                    for m in re.finditer(r"SGD\s*([\d,]+\.\d{2})", line_text):
                        amount = m.group(1)
                        start, end = m.span(0)
                        is_bold = False
                        cursor = 0
                        for s in spans:
                            s_text = s.get("text", "")
                            s_start = cursor
                            s_end = cursor + len(s_text)
                            cursor = s_end
                            if s_end > start and s_start < end:
                                if _is_span_bold(s):
                                    is_bold = True
                                    break
                        norm = norm_amount(amount)
                        result[norm] = result.get(norm, False) or is_bold
    finally:
        doc.close()
    return result


AMOUNT_RE = r"(?:SGD|S\$)?\s*([\d,]+\.\d{2})"


def extract_fields_from_pdf(text: str, excel_name: str) -> tuple[dict, dict]:
    """从 PDF 文本里抽各字段。

    返回 (fields, anomalies)：
    - fields: 各字段值，如 {pod_amount: "60,000.00", addr1: "...", ...}
    - anomalies: 检测到的 "0" 异常 {field_key: 描述}，
                 描述会直接作为 Issue 显示，且会跳过该字段的常规对比
    """
    out: dict[str, str] = {}
    anomalies: dict[str, str] = {}

    # POD 日期：第一段 "POD") dated 17 July 2025"
    m = re.search(
        r'POD[^)]{0,3}\)\s*dated\s+(\d{1,2}\s+\w+\s+\d{4})',
        text, re.IGNORECASE,
    )
    if m:
        out["pod_date"] = m.group(1)

    # POD amount：第一段 "in the amount of SGD X"
    m = re.search(
        r"in\s+the\s+amount\s+of\s+SGD\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE,
    )
    if m:
        out["pod_amount"] = m.group(1)

    # First Dividend：第二段 "you have been paid SGD X"
    m = re.search(
        r"you\s+have\s+been\s+paid\s+SGD\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE,
    )
    if m:
        out["first_div"] = m.group(1)

    # Admitted amount：第四段 "Amount of your POD admitted: SGD X"
    m = re.search(
        r"POD\s+admitted\s*:?\s*SGD\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE,
    )
    if m:
        out["admitted"] = m.group(1)

    # Final Dividend：第四段 "Final dividend payable to you: SGD X"
    m = re.search(
        r"Final\s+dividend\s+payable\s+to\s+you\s*:?\s*SGD\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE,
    )
    if m:
        out["final_div"] = m.group(1)

    # Bank Name
    m = re.search(r"Bank\s+Name\s*[:\-]\s*([^\n\r]+)", text, re.IGNORECASE)
    if m:
        out["bank_name"] = m.group(1).strip()

    # Bank Account Number
    m = re.search(r"Account\s+Number\s*[:\-]\s*([\d\-\s]+)", text, re.IGNORECASE)
    if m:
        out["bank_account"] = m.group(1).strip()

    # 地址块：定位收信人名字所在行，按位置取后续几行
    lines = [ln.strip() for ln in text.splitlines()]
    name_up = norm_text(excel_name)
    name_idx = -1
    for i, ln in enumerate(lines):
        if name_up and norm_text(ln) == name_up:
            name_idx = i
            break
    if name_idx >= 0:
        after = [ln for ln in lines[name_idx + 1: name_idx + 6] if ln]

        # 1) 异常检测：mail merge 把空字段渲染成了字面 "0"
        # 位置对应：第 0 行 = addr1, 第 1 行 = addr2, 第 2 行 = addr3 或 (postal + country),
        # 第 3 行 = country（如果分两行）
        pos_label = {0: "POD address 1", 1: "POD address 2",
                     2: "POD address 3", 3: "Country"}
        pos_key = {0: "addr1", 1: "addr2", 2: "addr3", 3: "country"}
        for i, line in enumerate(after[:4]):
            if line == "0":
                # 整行就是 "0"
                anomalies[pos_key[i]] = f"Remove 0 in {pos_label[i]}"
            elif re.match(r"^0\s+[A-Za-z]", line):
                # "0 SINGAPORE" / "0 MALAYSIA" 这种 0 + 国家名
                anomalies["country"] = "Remove 0 in Country"

        # 2) 按位置取值
        if len(after) >= 1:
            out["addr1"] = after[0]
        if len(after) >= 2:
            out["addr2"] = after[1]
        if len(after) >= 3:
            # 第三行通常是 "523859 SINGAPORE"
            third = after[2]
            m = re.match(r"(\d{4,6})\s+(.+)", third)
            if m:
                out["addr3"] = m.group(1)
                out["country"] = m.group(2)
            else:
                # 退而求其次：纯邮编 / 纯国家
                if re.fullmatch(r"\d{4,6}", third):
                    out["addr3"] = third
                else:
                    out["country"] = third
        # 有些信件邮编和国家分两行
        if "country" not in out and len(after) >= 4:
            out["country"] = after[3]

    # name 字段也填上，方便对比
    out.setdefault("name", excel_name)
    return out, anomalies


# ---------- 主流程 ----------

def main() -> None:
    parser = argparse.ArgumentParser(description="核对 PDF 信件与 Excel 数据")
    parser.add_argument("--excel", required=True, type=Path, help="Excel 文件路径")
    parser.add_argument("--sheet", default="expected_data", help="工作表名（默认 expected_data）")
    parser.add_argument("--pdf-dir", required=True, type=Path, help="PDF 所在文件夹")
    parser.add_argument("--out", default=Path("verification_report.xlsx"), type=Path,
                        help="输出 Excel 报告路径")
    parser.add_argument("--start-row", type=int, default=2,
                        help="起始 Excel 行号（1-based，含表头那行是 1；默认 2，即第一条数据）")
    parser.add_argument("--end-row", type=int, default=None,
                        help="结束 Excel 行号（含；默认到最后一行）")
    args = parser.parse_args()

    if not args.excel.exists():
        parser.error(f"Excel 不存在：{args.excel}")
    if not args.pdf_dir.exists():
        parser.error(f"PDF 文件夹不存在：{args.pdf_dir}")

    # 把 Excel 复制一份做报告
    shutil.copy(args.excel, args.out)
    # data_only=True：让 VLOOKUP 等公式返回缓存的计算结果，而不是公式字符串
    wb = load_workbook(args.out, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"❌ 工作表 {args.sheet!r} 不存在。可用工作表：{wb.sheetnames}")
        return
    ws = wb[args.sheet]

    # 找列号（1-based）。匹配时去掉列头前后空格，避免 "Bank Name " 这种意外。
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    header_map = {}
    for cell in header_row:
        if cell.value is None:
            continue
        key = str(cell.value).strip()
        header_map[key] = cell.column
    col_idx: dict[str, int] = {}
    missing_cols: list[str] = []
    for key, col_name in EXCEL_COLS.items():
        if col_name in header_map:
            col_idx[key] = header_map[col_name]
        else:
            missing_cols.append(col_name)
    if missing_cols:
        print(f"⚠ Excel 缺少列（这些字段会跳过）：{missing_cols}")
        print(f"   实际列头：{list(header_map.keys())}")

    # 在末尾加一列：问题摘要
    status_col = ws.max_column + 1
    ws.cell(row=1, column=status_col, value="Issues")

    field_mismatches: dict[str, int] = {}
    bold_mismatches: dict[str, int] = {}

    end_row = args.end_row or ws.max_row
    print(f"处理 Excel 行范围：{args.start_row} - {end_row}")

    # 在指定行范围里，把 Excel 名字归一化后建查找表
    name_to_row: dict[str, tuple] = {}
    for row in ws.iter_rows(min_row=args.start_row, max_row=end_row):
        name_cell = row[col_idx["name"] - 1] if "name" in col_idx else None
        if not name_cell or not name_cell.value:
            continue
        excel_name = str(name_cell.value).strip()
        name_to_row[norm_text(excel_name)] = (row, excel_name)

    # 扫每个 PDF，从内容里抓名字，再去 Excel 找对应行
    rows_with_issues = 0
    matched_count = 0
    unmatched_pdfs: list[tuple[str, str]] = []  # (pdf 文件名, 原因)

    pdf_files = sorted(args.pdf_dir.glob("*.pdf"))
    print(f"找到 PDF 文件：{len(pdf_files)} 个")

    for pdf_path in pdf_files:
        pdf_text = read_pdf_text(pdf_path)
        pdf_name = extract_recipient_name_from_text(pdf_text)

        if not pdf_name:
            unmatched_pdfs.append((pdf_path.name, "无法从 PDF 提取收信人名字"))
            continue

        match = name_to_row.get(norm_text(pdf_name))
        if not match:
            unmatched_pdfs.append((pdf_path.name, f"Excel {args.start_row}-{end_row} 行内找不到 {pdf_name}"))
            continue

        row, excel_name = match
        matched_count += 1

        pdf_fields, anomalies = extract_fields_from_pdf(pdf_text, excel_name)
        bold_map = get_amount_bold_map(pdf_path)

        issues: list[str] = []
        seen_anomalies: set[str] = set()
        for key, col in col_idx.items():
            excel_val = row[col - 1].value
            pdf_val = pdf_fields.get(key, "")
            kind = FIELD_KIND[key]

            # 优先处理 mail merge 的 "0" 异常，跳过该字段的常规对比
            if key in anomalies:
                msg = anomalies[key]
                if msg not in seen_anomalies:
                    issues.append(msg)
                    seen_anomalies.add(msg)
                field_mismatches[key] = field_mismatches.get(key, 0) + 1
                continue

            value_ok = normalize(excel_val, kind) == normalize(pdf_val, kind)

            if not value_ok:
                issues.append(f"{ISSUE_LABEL[key]} mismatch")
                field_mismatches[key] = field_mismatches.get(key, 0) + 1
                continue

            # 值正确，再看 SGD 粗体（仅对金额字段）
            if kind == "amount" and key in BOLD_EXPECTED:
                expected = BOLD_EXPECTED[key]
                actual = bold_map.get(norm_amount(pdf_val))
                if actual is None:
                    continue
                if actual != expected:
                    if expected:
                        issues.append(f"{ISSUE_LABEL[key]} SGD not bold")
                    else:
                        issues.append(f"{ISSUE_LABEL[key]} SGD should not be bold")
                    bold_mismatches[key] = bold_mismatches.get(key, 0) + 1

        if issues:
            ws.cell(row=row[0].row, column=status_col, value=", ".join(issues))
            rows_with_issues += 1
        else:
            ws.cell(row=row[0].row, column=status_col, value="-")

    wb.save(args.out)

    print(f"\n========== 核对完成 ==========")
    print(f"行范围：{args.start_row} - {end_row}")
    print(f"PDF 文件总数：{len(pdf_files)}")
    print(f"成功匹配到 Excel 行：{matched_count}")
    print(f"  其中有问题的：{rows_with_issues}")
    print(f"  完全匹配的：{matched_count - rows_with_issues}")
    print(f"没匹配上的 PDF：{len(unmatched_pdfs)}")
    if unmatched_pdfs:
        print(f"\n未匹配 PDF 列表（前 30 个）：")
        for fname, reason in unmatched_pdfs[:30]:
            print(f"  {fname}：{reason}")
    if field_mismatches:
        print(f"\n各字段值不匹配数：")
        for key, cnt in sorted(field_mismatches.items(), key=lambda x: -x[1]):
            print(f"  {EXCEL_COLS[key]}: {cnt}")
    if bold_mismatches:
        print(f"\n各字段 SGD 粗体不匹配数：")
        for key, cnt in sorted(bold_mismatches.items(), key=lambda x: -x[1]):
            expected = "粗体" if BOLD_EXPECTED[key] else "非粗体"
            print(f"  {EXCEL_COLS[key]} (应为{expected}): {cnt}")
    print(f"\n报告：{args.out}")
    print(f"  最后一列 Issues：'-' 表示全部正确；否则列出每个不匹配字段")


if __name__ == "__main__":
    main()
